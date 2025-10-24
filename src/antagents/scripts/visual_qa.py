import base64
import json
import mimetypes
import os
import uuid
from io import BytesIO

import PIL.Image
import requests
from dotenv import load_dotenv
from huggingface_hub import InferenceClient

from antagents import Tool, tool


load_dotenv(override=True)


def process_images_and_text(image_path, query, client):
    from transformers import AutoProcessor

    messages = [
        {
            "role": "user",
            "content": [
                {"type": "image"},
                {"type": "text", "text": query},
            ],
        },
    ]
    idefics_processor = AutoProcessor.from_pretrained("HuggingFaceM4/idefics2-8b-chatty")
    prompt_with_template = idefics_processor.apply_chat_template(messages, add_generation_prompt=True)

    # load images from local directory

    # encode images to strings which can be sent to the endpoint
    def encode_local_image(image_path):
        # load image
        image = PIL.Image.open(image_path).convert("RGB")

        # Convert the image to a base64 string
        buffer = BytesIO()
        image.save(buffer, format="JPEG")  # Use the appropriate format (e.g., JPEG, PNG)
        base64_image = base64.b64encode(buffer.getvalue()).decode("utf-8")

        # add string formatting required by the endpoint
        image_string = f"data:image/jpeg;base64,{base64_image}"

        return image_string

    image_string = encode_local_image(image_path)
    prompt_with_images = prompt_with_template.replace("<image>", "![]({}) ").format(image_string)

    payload = {
        "inputs": prompt_with_images,
        "parameters": {
            "return_full_text": False,
            "max_new_tokens": 200,
        },
    }

    return json.loads(client.post(json=payload).decode())[0]


# Function to encode the image
def encode_image(image_path):
    if image_path.startswith("http"):
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0"
        request_kwargs = {
            "headers": {"User-Agent": user_agent},
            "stream": True,
        }

        # Send a HTTP request to the URL
        response = requests.get(image_path, **request_kwargs)
        response.raise_for_status()
        content_type = response.headers.get("content-type", "")

        extension = mimetypes.guess_extension(content_type)
        if extension is None:
            extension = ".download"

        fname = str(uuid.uuid4()) + extension
        download_path = os.path.abspath(os.path.join("downloads", fname))

        with open(download_path, "wb") as fh:
            for chunk in response.iter_content(chunk_size=512):
                fh.write(chunk)

        image_path = download_path

    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")


def resize_image(image_path):
    img = PIL.Image.open(image_path)
    width, height = img.size
    img = img.resize((int(width / 2), int(height / 2)))
    new_image_path = f"resized_{image_path}"
    img.save(new_image_path)
    return new_image_path


class VisualQATool(Tool):
    name = "visualizer"
    description = "A tool that can answer questions about attached images."
    inputs = {
        "image_path": {
            "description": "The path to the image on which to answer the question",
            "type": "string",
        },
        "question": {"description": "the question to answer", "type": "string", "nullable": True},
    }
    output_type = "string"

    client = InferenceClient("HuggingFaceM4/idefics2-8b-chatty")

    def forward(self, image_path: str, question: str | None = None) -> str:
        output = ""
        add_note = False
        if not question:
            add_note = True
            question = "Please write a detailed caption for this image."
        try:
            output = process_images_and_text(image_path, question, self.client)
        except Exception as e:
            print(e)
            if "Payload Too Large" in str(e):
                new_image_path = resize_image(image_path)
                output = process_images_and_text(new_image_path, question, self.client)

        if add_note:
            output = (
                f"You did not provide a particular question, so here is a detailed caption for the image: {output}"
            )

        return output


def number_to_excel_column(n):
    """将数字转换为Excel列标格式（如1->A，27->AA）"""
    column = ""
    while n > 0:
        # 调整为0-25的范围（因为Excel列标无0）
        n -= 1
        # 计算当前字符对应的字母（0=A, 1=B...25=Z）
        column = chr(65 + n % 26) + column
        # 继续处理更高位
        n = n // 26
    return column


@tool
def visualizer(file_path: str, question: str | None = None) -> str:
    """A tool that can answer questions about attached images or Excel files (converted to images with cell colors).
    
    Args:
        file_path: The path to the image on which to answer the question. This should be a local path to downloaded image or Excel files.
        question: The question to answer.
    """
    import os
    add_note = False
    if not question:
        add_note = True
        question = "Please analyze this image and answer the question."
    if not isinstance(file_path, str):
        raise Exception("You must provide a valid file_path string")

    vlm_url = os.getenv("VLM_URL")
    vlm_api_key = os.getenv("VLM_API_KEY")
    
    if not vlm_url or not vlm_api_key:
        raise Exception("VLM_URL and VLM_API_KEY environment variables must be set")

    # 检查文件大小
    max_size = 20 * 1024 * 1024  # 20MB
    file_size = os.path.getsize(file_path)
    if file_size > max_size:
        raise Exception(f"File too large ({file_size/1024/1024:.2f}MB). Max allowed: 20MB")

    # 确定MIME类型
    import mimetypes
    mime_type, _ = mimetypes.guess_type(file_path)
    if not mime_type:
        mime_type = "application/octet-stream"
    
    # 处理Excel文件
    excel_mimes = {"application/vnd.ms-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    
    if mime_type in excel_mimes:
        try:
            # 导入必要的库
            import pandas as pd
            import matplotlib.pyplot as plt
            from io import BytesIO
            import base64
            import uuid
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            
            # 1. 读取Excel数据和样式（修复RGB对象处理）
            wb = load_workbook(file_path, data_only=False)
            ws = wb.active
            
            data = []
            cell_colors = []
            max_row = ws.max_row
            max_col = ws.max_column
            color_log = []
            
            for row in range(0, max_row + 1):
                data_row = []
                color_row = []
                
                if row == 0:
                    for col in range(0, max_col + 1):
                        if col == 0:
                            data_row.append("")
                        else:
                            data_row.append(number_to_excel_column(col))
                        hex_color = "#FFFFFF"  # 默认白色
                        color_row.append(hex_color)
                    data.append(data_row)
                    cell_colors.append(color_row)
                    continue
                
                for col in range(0, max_col + 1):
                    if col == 0:
                        data_row.append(str(row))
                        hex_color = "#F0F0F0"  # 默认浅灰色
                        color_row.append(hex_color)
                        continue
                    cell = ws.cell(row=row, column=col)
                    # 读取单元格值
                    cell_value = str(cell.value) if cell.value is not None else ""
                    data_row.append(cell_value)
                    
                    hex_color = "#FFFFFF"  # 默认白色
                    color_source = "default"
                    
                    # 改进的颜色读取逻辑
                    if cell.fill:
                        fill_obj = cell.fill
                        
                        # 处理StyleProxy
                        if hasattr(fill_obj, 'fill'):
                            fill_obj = fill_obj.fill
                            color_source = "StyleProxy_unwrapped"
                        
                        if isinstance(fill_obj, PatternFill):
                            # 处理前景色
                            if fill_obj.fgColor and fill_obj.fgColor.rgb:
                                rgb_value = fill_obj.fgColor.rgb
                                # 安全地处理RGB值
                                try:
                                    if hasattr(rgb_value, '__str__'):
                                        rgb_str = str(rgb_value).upper().replace(' ', '')
                                        # 确保是有效的16进制字符串
                                        if all(c in '0123456789ABCDEF' for c in rgb_str):
                                            if len(rgb_str) == 8 and rgb_str.startswith('FF'):
                                                hex_color = f"#{rgb_str[2:]}"
                                            elif len(rgb_str) == 6:
                                                hex_color = f"#{rgb_str}"
                                            else:
                                                hex_color = f"#{rgb_str[-6:]}"
                                        else:
                                            hex_color = "#FFFFFF"
                                    else:
                                        hex_color = "#FFFFFF"
                                except:
                                    hex_color = "#FFFFFF"
                                color_source = "fgColor"
                            # 处理背景色
                            elif fill_obj.bgColor and fill_obj.bgColor.rgb:
                                rgb_value = fill_obj.bgColor.rgb
                                try:
                                    if hasattr(rgb_value, '__str__'):
                                        rgb_str = str(rgb_value).upper().replace(' ', '')
                                        if all(c in '0123456789ABCDEF' for c in rgb_str):
                                            if len(rgb_str) == 8 and rgb_str.startswith('FF'):
                                                hex_color = f"#{rgb_str[2:]}"
                                            elif len(rgb_str) == 6:
                                                hex_color = f"#{rgb_str}"
                                            else:
                                                hex_color = f"#{rgb_str[-6:]}"
                                        else:
                                            hex_color = "#FFFFFF"
                                    else:
                                        hex_color = "#FFFFFF"
                                except:
                                    hex_color = "#FFFFFF"
                                color_source = "bgColor"
                            else:
                                color_source = "PatternFill_no_color"
                        else:
                            # 对于其他填充类型，尝试通用方法
                            color_source = f"other_fill_{type(fill_obj).__name__}"
                            # 尝试从各种颜色属性中提取
                            for attr in ['fgColor', 'bgColor', 'start_color', 'end_color']:
                                if hasattr(fill_obj, attr):
                                    color_obj = getattr(fill_obj, attr)
                                    if color_obj and hasattr(color_obj, 'rgb') and color_obj.rgb:
                                        rgb_value = color_obj.rgb
                                        try:
                                            if hasattr(rgb_value, '__str__'):
                                                rgb_str = str(rgb_value).upper().replace(' ', '')
                                                if all(c in '0123456789ABCDEF' for c in rgb_str):
                                                    if len(rgb_str) == 8 and rgb_str.startswith('FF'):
                                                        hex_color = f"#{rgb_str[2:]}"
                                                    elif len(rgb_str) == 6:
                                                        hex_color = f"#{rgb_str}"
                                                    else:
                                                        hex_color = f"#{rgb_str[-6:]}"
                                                else:
                                                    hex_color = "#FFFFFF"
                                            else:
                                                hex_color = "#FFFFFF"
                                        except:
                                            hex_color = "#FFFFFF"
                                        color_source = f"other_{attr}"
                                        break
                    else:
                        color_source = "no_fill"
                    
                    color_row.append(hex_color)
                    color_log.append(f"Cell ({row},{col}): value='{cell_value}', color={hex_color}, source={color_source}")
                
                data.append(data_row)
                cell_colors.append(color_row)
            
            # 打印日志
            print("Color reading log:")
            for log_entry in color_log[:20]:
                print(log_entry)
            if len(color_log) > 20:
                print(f"... and {len(color_log)-20} more entries")
            
            # 2. 绘图参数
            num_rows = len(data)
            num_cols = len(data[0]) if num_rows > 0 else 0
            fig_width = min(25, max(10, num_cols * 2))
            fig_height = min(35, max(8, num_rows * 0.8))
            font_size = max(8, 12 - (num_rows // 10))
            
            # 3. 字体配置
            plt.rcParams["font.family"] = ["SimHei", "Microsoft YaHei", "Arial", "sans-serif"]
            plt.rcParams["axes.unicode_minus"] = False
            
            # 4. 创建表格
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            ax.axis('tight')
            ax.axis('off')
            table = ax.table(
                cellText=data,
                cellLoc='center',
                loc='center',
                bbox=[0, 0, 1, 1],
                edges='closed'
            )
            table.auto_set_font_size(False)
            table.set_fontsize(font_size)
            table.scale(1.5, 2.0)
            
            # 5. 应用颜色
            for row_idx in range(num_rows):
                for col_idx in range(num_cols):
                    cell_key = (row_idx, col_idx)
                    if cell_key in table.get_celld():
                        cell = table.get_celld()[cell_key]
                        cell_color = cell_colors[row_idx][col_idx] if (row_idx < len(cell_colors) and col_idx < len(cell_colors[row_idx])) else "#FFFFFF"
                        # 确保颜色格式正确
                        if not cell_color.startswith('#') or len(cell_color) != 7:
                            cell_color = "#FFFFFF"
                        cell.set_facecolor(cell_color)
                        cell.set_edgecolor('#333333')
                        cell.set_linewidth(2)
                        # 文字颜色适配
                        try:
                            color_hex = cell_color.lstrip('#')
                            if len(color_hex) == 6:
                                r, g, b = int(color_hex[0:2],16), int(color_hex[2:4],16), int(color_hex[4:6],16)
                                text_color = 'black' if (0.299*r + 0.587*g + 0.114*b)/255 > 0.5 else 'white'
                                cell.get_text().set_color(text_color)
                                cell.get_text().set_weight('bold')
                                cell.get_text().set_size(font_size)
                        except:
                            cell.get_text().set_color('black')
                            cell.get_text().set_weight('bold')
                            cell.get_text().set_size(font_size)
            
            # 6. 保存图片
            buf = BytesIO()
            plt.savefig(buf, format='png', bbox_inches='tight', dpi=300, facecolor='white')
            buf.seek(0)
            # 保存本地调试图
            base_filename = os.path.splitext(os.path.basename(file_path))[0]
            unique_id = uuid.uuid4().hex[:8]
            debug_image_path = f"{base_filename}_debug_{unique_id}.png"
            with open(debug_image_path, "wb") as f:
                f.write(buf.getvalue())
            print(f"Debug image saved to: {os.path.abspath(debug_image_path)}")
            # 转base64
            buf.seek(0)
            base64_data = base64.b64encode(buf.read()).decode('utf-8')
            mime_type = "image/png"
            buf.close()
            plt.close()
            
        except Exception as e:
            raise Exception(f"Error converting Excel to image: {str(e)}")
    else:
        # 处理非Excel文件
        try:
            import base64
            with open(file_path, "rb") as file:
                base64_data = base64.b64encode(file.read()).decode('utf-8')
        except Exception as e:
            raise Exception(f"Error reading file: {str(e)}")
    
    # 构建请求
    content = [{"type": "text", "text": question}]
    image_mimes = {"image/png", "image/jpeg", "image/webp", "image/gif"}
    if mime_type in image_mimes:
        content.append({"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{base64_data}"}})
    else:
        try:
            with open(file_path, "rb") as file:
                text_data = file.read().decode("utf-8", errors="ignore")[:4000]
        except:
            text_data = "<无法解析为文本>"
        content.append({"type": "input_text", "text": text_data})
    
    # 发送请求
    try:
        import requests
        import json
        headers = {"Content-Type": "application/json", "Authorization": f"{vlm_api_key}"}
        payload = {"model": os.getenv("VLM_MODEL_ID"), "messages": [{"role": "user", "content": content}], "max_tokens": 1000}
        response = requests.post(vlm_url, headers=headers, json=payload, timeout=60)
        response.raise_for_status()
        result = response.json()
    except Exception as e:
        raise Exception(f"API request failed: {str(e)}")
    
    if "choices" in result:
        output = result["choices"][0]["message"]["content"]
    else:
        raise Exception(f"Unexpected response format: {result}")

    if add_note:
        output = f"Analysis: {output}"
    return output